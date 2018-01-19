#! /usr/bin/env python3
# # -*- coding: utf-8 -*-
"""WeArePythonistas - Part 0: What do we want to see"""
# We read in an xlsx and crunch the numbers to get a chart

from openpyxl import load_workbook
from operator import itemgetter
import matplotlib.pyplot as plt

# I put the file on my machine so I have a clear privacy separation
file='/Users/stefansteinbauer/Documents/onmachinedata/report.xlsx'
wsName='Attendees'
wb = load_workbook(file)
ws = wb[wsName]

# column 22 for languages, 31 for countries
colVal=22
# when you have multiple operations you might want to have only a subset, like the
# first 1000 rows
# rowmax=1000
rowmax=ws.max_row
# read values from column 22 (languages) into list
# not very pythonic, but here you also see the stripping and uppercasing to
# get through the masses of different ways to write Javascript
interests=list()
for i in range(2, rowmax):
    cellval = str(ws.cell(row=i, column=colVal).value)
    cellval = cellval.replace(' ','')
    cellval = cellval.upper()
# and here I look at people filling in their languageS
    interests.extend(cellval.split(","))
#counting the languages while removing the duplicates
lang2={x:interests.count(x) for x in interests}
# Getting the Top 10 or more
lang = dict(sorted(lang2.items(), key=itemgetter(1), reverse=True)[:10])


# let's print that - xkcd style
with plt.xkcd():
    # removing the right and top spines and shifting to the right to unclip the labels
    fig = plt.figure()
    ax = fig.add_subplot(1, 1, 1)
    ax.spines['right'].set_color('none')
    ax.spines['top'].set_color('none')
    plt.gcf().subplots_adjust(left=0.3)
    plt.title("WHAT WE WANT")
    plt.rcParams.update({'figure.autolayout': True})
    plt.barh(range(len(lang)),list(lang.values()),align='center')
    # display values next to bars
    #for i, v in enumerate(list(lang.values())):
    #    ax.text(v + 4, i-.35, str(v), fontweight='bold')
    plt.yticks(range(len(lang)),list(lang.keys()))
    # plt.show()
    fig.savefig('report'+str(colVal)+'.png')
